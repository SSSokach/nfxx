from fastapi import APIRouter, Depends, UploadFile, File as FastAPIFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import List, Optional
from app import SessionLocal
from models import File, FileCollectionTask, FileCollectionItem, ChatTodoItem, Message, User
from services.excel_analyzer import read_excel_template
from datetime import datetime
import os
import io

router = APIRouter()

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@router.get("/{user_id}")
def get_files(user_id: int, db: Session = Depends(get_db)):
    files = db.query(File).filter(File.owner_id == user_id).all()
    return [{"id": f.id, "name": f.name, "file_type": f.file_type, "created_at": f.created_at.isoformat()} for f in files]

@router.get("/content/{file_id}")
def get_file_content(file_id: int, db: Session = Depends(get_db)):
    file = db.query(File).filter(File.id == file_id).first()
    if file:
        return {"id": file.id, "name": file.name, "content": file.content, "file_type": file.file_type}
    return {"error": "File not found"}

@router.get("/by-name/{file_name}")
def get_file_by_name(file_name: str, db: Session = Depends(get_db)):
    """通过文件名查找文件。"""
    file = db.query(File).filter(File.name == file_name).order_by(File.created_at.desc()).first()
    if file:
        return {"id": file.id, "name": file.name, "content": file.content, "file_type": file.file_type}
    return {"error": "File not found"}

@router.post("/save")
def save_file(user_id: int, name: str, content: str, file_type: str = "markdown", db: Session = Depends(get_db)):
    file = File(
        name=name,
        content=content,
        file_type=file_type,
        owner_id=user_id,
        created_at=datetime.utcnow()
    )
    db.add(file)
    db.commit()
    db.refresh(file)
    return {"id": file.id, "name": file.name, "created_at": file.created_at.isoformat()}

@router.put("/update/{file_id}")
def update_file(file_id: int, content: str, db: Session = Depends(get_db)):
    file = db.query(File).filter(File.id == file_id).first()
    if file:
        file.content = content
        db.commit()
        return {"message": "File updated successfully"}
    return {"error": "File not found"}

@router.post("/upload/{user_id}")
async def upload_file(user_id: int, file: UploadFile = FastAPIFile(...), db: Session = Depends(get_db)):
    """上传本地文件，保存内容到数据库，同时保存物理文件到uploads目录。"""
    content = await file.read()
    text_content = ""
    try:
        text_content = content.decode("utf-8")
    except UnicodeDecodeError:
        text_content = f"[二进制文件，大小: {len(content)} 字节]"

    # Determine file type
    filename = file.filename or "unnamed"
    ext = os.path.splitext(filename)[1].lower().lstrip(".")
    if ext in ("md", "markdown"):
        file_type = "markdown"
    elif ext in ("txt",):
        file_type = "text"
    elif ext in ("json",):
        file_type = "json"
    elif ext in ("py", "js", "ts", "java", "go", "rs", "c", "cpp", "html", "css", "vue"):
        file_type = "code"
    else:
        file_type = ext or "file"

    # Save to database
    file_obj = File(
        name=filename,
        content=text_content,
        file_type=file_type,
        owner_id=user_id,
        created_at=datetime.utcnow()
    )
    db.add(file_obj)
    db.commit()
    db.refresh(file_obj)

    # Save physical file
    phys_path = os.path.join(UPLOAD_DIR, f"{file_obj.id}_{filename}")
    with open(phys_path, "wb") as f:
        f.write(content)

    return {
        "id": file_obj.id,
        "name": file_obj.name,
        "file_type": file_obj.file_type,
        "size": len(content),
        "created_at": file_obj.created_at.isoformat()
    }

@router.get("/download/{file_id}")
def download_file(file_id: int, db: Session = Depends(get_db)):
    """下载文件，优先返回物理文件，没有则用数据库内容生成。"""
    file = db.query(File).filter(File.id == file_id).first()
    if not file:
        return {"error": "File not found"}

    phys_path = os.path.join(UPLOAD_DIR, f"{file.id}_{file.name}")
    if os.path.exists(phys_path):
        with open(phys_path, "rb") as f:
            content = f.read()
    else:
        content = file.content.encode("utf-8")

    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{file.name}"}
    )

@router.delete("/delete/{file_id}")
def delete_file(file_id: int, db: Session = Depends(get_db)):
    """删除文件。"""
    file = db.query(File).filter(File.id == file_id).first()
    if not file:
        return {"error": "File not found"}

    # Delete physical file if exists
    phys_path = os.path.join(UPLOAD_DIR, f"{file.id}_{file.name}")
    if os.path.exists(phys_path):
        os.remove(phys_path)

    db.delete(file)
    db.commit()
    return {"message": "File deleted successfully"}


@router.get("/excel/{file_id}")
def get_excel_content(file_id: int, db: Session = Depends(get_db)):
    """读取 Excel 文件，返回 JSON 格式的表格数据用于在线预览。"""
    file = db.query(File).filter(File.id == file_id).first()
    if not file:
        return {"error": "File not found"}

    phys_path = os.path.join(UPLOAD_DIR, f"{file.id}_{file.name}")
    if not os.path.exists(phys_path):
        return {"error": "Physical file not found"}

    try:
        from openpyxl import load_workbook
        wb = load_workbook(phys_path, data_only=True)
    except Exception as e:
        return {"error": f"读取Excel失败: {e}"}

    sheets = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=False):
            cells = []
            for cell in row:
                val = cell.value
                if val is None:
                    cells.append("")
                elif isinstance(val, float) and val.is_integer():
                    cells.append(str(int(val)))
                else:
                    cells.append(str(val))
            rows.append(cells)

        # 识别空行，截断
        last_non_empty = 0
        for i, row in enumerate(rows):
            if any(c.strip() for c in row if isinstance(c, str)):
                last_non_empty = i
        rows = rows[:last_non_empty + 1]

        sheets.append({
            "name": ws.title,
            "rows": rows,
            "max_rows": len(rows),
            "max_cols": max((len(r) for r in rows), default=0),
        })

    return {
        "file_id": file.id,
        "file_name": file.name,
        "sheets": sheets,
    }


class CellChange(BaseModel):
    row: int
    col: int
    value: str


class ExcelSaveRequest(BaseModel):
    sheet_index: int = 0
    changes: List[CellChange] = []
    user_id: Optional[int] = None


@router.put("/excel/{file_id}")
def save_excel_content(file_id: int, req: ExcelSaveRequest, db: Session = Depends(get_db)):
    """保存 Excel 单元格修改，写回物理文件。
    同时检查文件收集任务，更新填写人待办状态。
    """
    file = db.query(File).filter(File.id == file_id).first()
    if not file:
        return {"error": "File not found"}

    phys_path = os.path.join(UPLOAD_DIR, f"{file.id}_{file.name}")
    if not os.path.exists(phys_path):
        return {"error": "Physical file not found"}

    try:
        from openpyxl import load_workbook
        wb = load_workbook(phys_path)
    except Exception as e:
        return {"error": f"读取Excel失败: {e}"}

    # 应用单元格修改
    sheets = wb.worksheets
    if req.sheet_index < 0 or req.sheet_index >= len(sheets):
        return {"error": "无效的工作表索引"}

    ws = sheets[req.sheet_index]
    for change in req.changes:
        # row/col 从0开始（前端），Excel从1开始
        # 第0行是表头，不可编辑
        if change.row == 0:
            continue
        ws.cell(row=change.row + 1, column=change.col + 1, value=change.value)

    wb.save(phys_path)

    # === 更新文件收集任务和待办 ===
    result = {"saved": True, "changes_count": len(req.changes)}

    if req.user_id:
        # 查找关联的文件收集任务
        tasks = db.query(FileCollectionTask).filter(
            FileCollectionTask.file_name == file.name
        ).all()

        for task in tasks:
            # 重新读取 Excel 检查填写状态
            analysis = read_excel_template(file.id, file.name)
            if "error" in analysis:
                continue

            items = db.query(FileCollectionItem).filter(
                FileCollectionItem.task_id == task.id
            ).all()

            newly_submitted = 0
            for item in items:
                person = next(
                    (p for p in analysis["people"] if p["name"] == item.assignee_name),
                    None,
                )
                if person and person["filled"] and item.status != "submitted":
                    item.status = "submitted"
                    item.submitted_at = datetime.utcnow()
                    newly_submitted += 1

                    # 更新填写人待办为已完成
                    assignee_todos = db.query(ChatTodoItem).filter(
                        ChatTodoItem.user_id == item.assignee_user_id,
                        ChatTodoItem.source_id == task.source_message_id,
                        ChatTodoItem.content.like(f"%[填写文件]%{task.file_name}%"),
                    ).all()
                    for at in assignee_todos:
                        if at.status == "pending":
                            at.status = "completed"
                            at.completed_at = datetime.utcnow()
                            at.updated_at = datetime.utcnow()

            # 更新进度
            submitted_count = sum(1 for i in items if i.status == "submitted")
            total_count = len(items)

            # 更新发起人待办
            initiator_todos = db.query(ChatTodoItem).filter(
                ChatTodoItem.user_id == task.initiator_user_id,
                ChatTodoItem.source_id == task.source_message_id,
                ChatTodoItem.content.like(f"%[文件收集]%{task.file_name}%"),
            ).all()
            for it in initiator_todos:
                if submitted_count == total_count and total_count > 0:
                    it.content = f"[文件收集] {task.file_name} - 全部已提交（{submitted_count}/{total_count}）✓"
                    task.status = "completed"
                else:
                    it.content = f"[文件收集] {task.file_name} - 等待 {total_count} 人提交（{submitted_count}/{total_count}）"
                it.updated_at = datetime.utcnow()

            task.updated_at = datetime.utcnow()
            result["task_id"] = task.id
            result["submitted"] = submitted_count
            result["total"] = total_count
            result["newly_submitted"] = newly_submitted

    db.commit()
    return result
