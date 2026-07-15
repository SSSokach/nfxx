"""Excel 文件分析服务：读取 Excel 模板，提取待填写人名单和填写状态。"""

import os
from openpyxl import load_workbook
from datetime import datetime

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "uploads")


def _get_excel_path(file_id: int, file_name: str):
    """获取 Excel 文件的物理路径。"""
    phys_path = os.path.join(UPLOAD_DIR, f"{file_id}_{file_name}")
    if os.path.exists(phys_path):
        return phys_path
    return None


def read_excel_template(file_id: int, file_name: str):
    """读取 Excel 模板，提取：
    - 文件名
    - 表头列名
    - 待填写人名单（从"姓名"列）
    - 每人填写状态（已填写/未填写，根据其他列是否有内容判断）
    
    返回:
    {
        "file_name": str,
        "headers": [str],
        "people": [
            {
                "name": str,
                "row_index": int,
                "filled": bool,
                "filled_columns": [str],  # 已填写的列名
                "empty_columns": [str],   # 未填写的列名
            }
        ],
        "total": int,
        "filled_count": int,
        "unfilled_count": int,
    }
    """
    path = _get_excel_path(file_id, file_name)
    if not path:
        return {"error": f"找不到Excel文件: {file_name}"}

    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        return {"error": f"读取Excel失败: {e}"}

    ws = wb.active
    if not ws or ws.max_row < 2:
        return {"error": "Excel文件为空或只有表头"}

    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value is not None:
            headers.append(str(cell.value).strip())
        else:
            break

    if not headers:
        return {"error": "Excel没有表头行"}

    # 找到"姓名"列
    name_col_idx = None
    for i, h in enumerate(headers):
        if h in ("姓名", "名字", "name", "Name", "NAME", "员工姓名"):
            name_col_idx = i
            break

    if name_col_idx is None:
        # 如果没有姓名列，用第一列作为姓名
        name_col_idx = 0

    # 需要检查填写的列（排除姓名列）
    check_cols = [i for i in range(len(headers)) if i != name_col_idx]

    # 读取每个人的填写状态
    people = []
    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]
        name_cell = row[name_col_idx].value if name_col_idx < len(row) else None
        if not name_cell:
            continue

        name = str(name_cell).strip()
        filled_cols = []
        empty_cols = []

        for col_idx in check_cols:
            if col_idx >= len(row):
                continue
            cell_val = row[col_idx].value
            col_name = headers[col_idx]
            if cell_val is not None and str(cell_val).strip() != "":
                filled_cols.append(col_name)
            else:
                empty_cols.append(col_name)

        is_filled = len(empty_cols) == 0 if check_cols else True
        if not check_cols:
            is_filled = True

        people.append({
            "name": name,
            "row_index": row_idx,
            "filled": is_filled,
            "filled_columns": filled_cols,
            "empty_columns": empty_cols,
        })

    total = len(people)
    filled_count = sum(1 for p in people if p["filled"])
    unfilled_count = total - filled_count

    return {
        "file_name": file_name,
        "headers": headers,
        "people": people,
        "total": total,
        "filled_count": filled_count,
        "unfilled_count": unfilled_count,
    }


def analyze_excel_with_ai(file_id: int, file_name: str, message_content: str = ""):
    """AI 分析 Excel 文件 + 消息内容，判断是否为文件收集请求。
    
    返回:
    {
        "is_collection_request": bool,
        "file_name": str,
        "description": str,
        "deadline": str | None,
        "assignees": [str],  # 待填写人名单
        "unfilled": [str],   # 未填写人名单
    }
    """
    # 1. 读取 Excel 模板
    result = read_excel_template(file_id, file_name)
    if "error" in result:
        return result

    # 2. 检测消息是否为文件收集请求
    collect_keywords = ["填写", "提交", "收集", "汇总", "填报", "收一下", "交一下", "填写文件", "交表"]
    file_keywords = ["文件", "文档", "表格", "周报", "月报", "报告", "清单", "模板", "excel", "Excel"]
    
    is_collect = any(kw in message_content for kw in collect_keywords)
    has_file_ref = any(kw in message_content for kw in file_keywords) or True  # 有文件本身就算

    # 3. 提取待填写人名单（从 Excel 的姓名列）
    assignees = [p["name"] for p in result["people"]]
    unfilled = [p["name"] for p in result["people"] if not p["filled"]]

    # 4. 提取截止时间
    import re
    from datetime import date, timedelta
    deadline = None
    m = re.search(r'(\d{1,2})月(\d{1,2})号?', message_content)
    if m:
        month, day = int(m.group(1)), int(m.group(2))
        try:
            deadline = date(date.today().year, month, day).isoformat()
        except ValueError:
            pass
    if not deadline:
        if "今天" in message_content:
            deadline = date.today().isoformat()
        elif "明天" in message_content:
            deadline = (date.today() + timedelta(days=1)).isoformat()
        elif "本周五" in message_content or "周五" in message_content:
            today = date.today()
            days_ahead = (4 - today.weekday() + 7) % 7
            deadline = (today + timedelta(days=days_ahead)).isoformat()
        elif "下周一" in message_content or "周一" in message_content:
            today = date.today()
            days_ahead = (0 - today.weekday() + 7) % 7
            deadline = (today + timedelta(days=days_ahead)).isoformat()
        elif "下周" in message_content:
            deadline = (date.today() + timedelta(days=7)).isoformat()

    return {
        "is_collection_request": is_collect or len(assignees) > 0,
        "file_name": file_name,
        "headers": result["headers"],
        "description": message_content[:200] if message_content else f"填写{file_name}",
        "deadline": deadline,
        "assignees": assignees,
        "unfilled": unfilled,
        "total": result["total"],
        "filled_count": result["filled_count"],
        "unfilled_count": result["unfilled_count"],
        "people_detail": result["people"],
    }
